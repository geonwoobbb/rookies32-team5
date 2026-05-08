import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime


EXCEL_FILE = "Pokemon_Card_Price_Report.xlsx"


def create_excel_report(crawling_data, pokemon_name):
    if not crawling_data:
        print("엑셀에 저장할 데이터가 없습니다.")
        return None

    now = datetime.now().strftime("%Y%m%d_%H%M")

    try:
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Price_List"

            header = ["No", "Pokemon", "Title", "Price", "Time"]
            ws.append(header)

            header_fill = PatternFill(
                start_color="FF00FF",
                fill_type="solid"
            )

            header_font = Font(
                color="FFFFFF",
                bold=True
            )

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill

        start_no = ws.max_row

        for i, item in enumerate(crawling_data, 1):
            row = [
                start_no + i - 1,
                pokemon_name,
                item["title"],
                item["price"],
                now
            ]

            ws.append(row)

        wb.save(EXCEL_FILE)

        print(f"엑셀 생성 완료: {EXCEL_FILE}")
        print(f"저장 위치: {os.path.abspath(EXCEL_FILE)}")

        return EXCEL_FILE

    except PermissionError:
        print("엑셀 파일이 열려 있어서 저장할 수 없습니다.")
        print("Pokemon_Card_Price_Report.xlsx 파일을 닫고 다시 실행하세요.")
        return None