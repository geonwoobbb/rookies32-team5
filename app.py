from flask import Flask, render_template, jsonify, request, send_file
import json, io, os, requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime

app = Flask(__name__)

FAVORITES_FILE = "favorites.json"
MAX_FAVORITES = 5

KOREAN_NAMES = {
    "bulbasaur":"이상해씨","ivysaur":"이상해풀","venusaur":"이상해꽃",
    "charmander":"파이리","charmeleon":"리자드","charizard":"리자몽",
    "squirtle":"꼬부기","wartortle":"어니부기","blastoise":"거북왕",
    "caterpie":"캐터피","metapod":"단데기","butterfree":"버터플",
    "weedle":"뿔충이","kakuna":"딱구리","beedrill":"독침붕",
    "pidgey":"구구","pidgeotto":"피죤","pidgeot":"피죤투",
    "rattata":"꼬렛","raticate":"레트라",
    "spearow":"깨비참","fearow":"폴켓",
    "ekans":"아보","arbok":"아보크",
    "pikachu":"피카츄","raichu":"라이츄",
    "sandshrew":"모래두지","sandslash":"고지",
    "nidoran-f":"니드런♀","nidorina":"니드리나","nidoqueen":"니드퀸",
    "nidoran-m":"니드런♂","nidorino":"니드리노","nidoking":"니드킹",
    "clefairy":"삐삐","clefable":"픽시",
    "vulpix":"식스테일","ninetales":"나인테일",
    "jigglypuff":"푸린","wigglytuff":"푸크린",
    "zubat":"주뱃","golbat":"골뱃",
    "oddish":"뚜벅초","gloom":"냄새꼬","vileplume":"라플레시아",
    "paras":"파라스","parasect":"파라섹트",
    "venonat":"콘팡","venomoth":"도나리",
    "diglett":"디그다","dugtrio":"닥트리오",
    "meowth":"나옹","persian":"페르시온",
    "psyduck":"고라파덕","golduck":"골덕",
    "mankey":"망키","primeape":"성원숭",
    "growlithe":"가디","arcanine":"윈디",
    "poliwag":"발챙이","poliwhirl":"슈륙챙이","poliwrath":"강챙이",
    "abra":"케이시","kadabra":"윤겔라","alakazam":"후딘",
    "machop":"알통몬","machoke":"근육몬","machamp":"괴력몬",
    "bellsprout":"모다피","weepinbell":"우츠동","victreebel":"우츠보트",
    "tentacool":"왕눈해","tentacruel":"독파리",
    "geodude":"꼬마돌","graveler":"돌돌이","golem":"golem",
    "ponyta":"포니타","rapidash":"날쌩마",
    "slowpoke":"야돈","slowbro":"야부",
    "magnemite":"코일","magneton":"레어코일",
    "farfetchd":"파오리","doduo":"두두","dodrio":"두트리오",
    "seel":"쥬쥬","dewgong":"쥬레곤",
    "grimer":"질퍽이","muk":"질뻐기",
    "shellder":"셀러","cloyster":"파르셀",
    "gastly":"고오스","haunter":"고우스트","gengar":"팬텀",
    "onix":"롱스톤","drowzee":"슬리프","hypno":"슬리퍼",
    "krabby":"크랩","kingler":"킹크랩",
    "voltorb":"찌리리공","electrode":"붐볼",
    "exeggcute":"아라리","exeggutor":"나시",
    "cubone":"탕구리","marowak":"마로왁",
    "hitmonlee":"시라소몬","hitmonchan":"에비몬",
    "lickitung":"내루미","koffing":"또가스","weezing":"독독이",
    "rhyhorn":"뿔카노","rhydon":"강철톤",
    "chansey":"럭키","tangela":"덩쿠리","kangaskhan":"캥카",
    "horsea":"쏘드라","seadra":"시드라",
    "goldeen":"콘치","seaking":"왕콘치",
    "staryu":"별가사리","starmie":"아쿠스타",
    "mr-mime":"마임맨","scyther":"스라크","jynx":"루주라",
    "electabuzz":"에레브","magmar":"마그마",
    "pinsir":"쁘사이저","tauros":"켄타로스",
    "magikarp":"잉어킹","gyarados":"갸라도스",
    "lapras":"라프라스","ditto":"메타몽",
    "eevee":"이브이","vaporeon":"샤미드","jolteon":"쥬피썬더","flareon":"부스터",
    "porygon":"폴리곤",
    "omanyte":"암나이트","omastar":"암스타",
    "kabuto":"투구","kabutops":"투구푸스",
    "aerodactyl":"프테라","snorlax":"잠만보",
    "articuno":"프리져","zapdos":"썬더","moltres":"파이어",
    "dratini":"미니룡","dragonair":"신룡","dragonite":"망나뇽",
    "mewtwo":"뮤츠","mew":"뮤",
}

TYPE_KO = {
    "grass":"풀","fire":"불꽃","water":"물","bug":"벌레",
    "normal":"노말","poison":"독","electric":"전기","ground":"땅",
    "fairy":"페어리","fighting":"격투","psychic":"에스퍼","rock":"바위",
    "ghost":"고스트","ice":"얼음","dragon":"드래곤","steel":"강철",
    "flying":"비행","dark":"악",
}
TYPE_COLORS = {
    "grass":"#78C850","fire":"#F08030","water":"#6890F0",
    "bug":"#A8B820","normal":"#A8A878","poison":"#A040A0",
    "electric":"#F8D030","ground":"#E0C068","fairy":"#EE99AC",
    "fighting":"#C03028","psychic":"#F85888","rock":"#B8A038",
    "ghost":"#705898","ice":"#98D8D8","dragon":"#7038F8",
    "steel":"#B8B8D0","flying":"#A890F0","dark":"#705848",
}

# ── 유틸 ──────────────────────────────────────────
def load_cache():
    with open("pokemon_cache.json","r",encoding="utf-8") as f:
        return json.load(f)

def get_korean_name(eng):
    return KOREAN_NAMES.get(eng, eng)

def load_favorites():
    if not os.path.exists(FAVORITES_FILE):
        return []
    with open(FAVORITES_FILE,"r",encoding="utf-8") as f:
        return json.load(f)

def save_favorites(favs):
    with open(FAVORITES_FILE,"w",encoding="utf-8") as f:
        json.dump(favs, f, ensure_ascii=False)

def crawling(keyword, limit=5):
    search_keyword = keyword + " 카드"
    url = f"https://web.joongna.com/search/{search_keyword}"
    headers = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'}
    try:
        response = requests.get(url, headers=headers, timeout=10)
        soup = BeautifulSoup(response.text,'html.parser')
        items = soup.select("div.z-auto")
        result = []
        for item in items:
            title = item.select_one(".whitespace-pre-line")
            price = item.select_one(".text-18")
            if title and price:
                result.append({
                    "title": title.get_text(strip=True),
                    "price": price.get_text(strip=True),
                })
            if len(result) >= limit:
                break
        return result
    except:
        return []

# ── 라우트 ────────────────────────────────────────
@app.route("/")
def index():
    cache = load_cache()
    pokemon_data = cache["pokemon"]
    favorites = load_favorites()
    all_types = sorted(set(t for p in pokemon_data.values() for t in p["types"]))
    pokemon_list = []
    for eng, data in pokemon_data.items():
        pokemon_list.append({
            "id": eng,
            "name": get_korean_name(eng),
            "eng_name": eng,
            "types": data["types"],
            "image": data["image"],
            "favorited": eng in favorites,
        })
    return render_template("index.html",
        pokemon_list=pokemon_list,
        all_types=all_types,
        type_ko=TYPE_KO,
        type_colors=TYPE_COLORS,
        favorites=favorites,
        max_favorites=MAX_FAVORITES,
    )

@app.route("/pokemon/<name>")
def detail(name):
    cache = load_cache()
    pokemon_data = cache["pokemon"]
    if name not in pokemon_data:
        return "포켓몬을 찾을 수 없습니다.", 404
    p = pokemon_data[name]
    favorites = load_favorites()
    evolution_chain = []
    for evo in p.get("evolution", []):
        evolution_chain.append({
            "eng": evo,
            "kor": get_korean_name(evo),
            "image": pokemon_data[evo]["image"] if evo in pokemon_data else "",
        })
    return render_template("detail.html",
        pokemon={
            "id": name,
            "name": get_korean_name(name),
            "eng_name": name,
            "types": p["types"],
            "moves": p.get("moves", []),
            "evolution": evolution_chain,
            "image": p["image"],
        },
        type_ko=TYPE_KO,
        type_colors=TYPE_COLORS,
        favorited=(name in favorites),
        fav_count=len(favorites),
        max_favorites=MAX_FAVORITES,
    )

@app.route("/favorites")
def favorites_page():
    cache = load_cache()
    pokemon_data = cache["pokemon"]
    favorites = load_favorites()
    fav_list = []
    for eng in favorites:
        if eng in pokemon_data:
            data = pokemon_data[eng]
            fav_list.append({
                "id": eng,
                "name": get_korean_name(eng),
                "eng_name": eng,
                "types": data["types"],
                "image": data["image"],
            })
    return render_template("favorites.html",
        fav_list=fav_list,
        type_ko=TYPE_KO,
        type_colors=TYPE_COLORS,
        max_favorites=MAX_FAVORITES,
    )

# ── API ───────────────────────────────────────────
@app.route("/api/favorite", methods=["POST"])
def toggle_favorite():
    data = request.get_json()
    name = data.get("name","")
    favorites = load_favorites()
    if name in favorites:
        favorites.remove(name)
        save_favorites(favorites)
        return jsonify({"status":"removed","count":len(favorites)})
    if len(favorites) >= MAX_FAVORITES:
        return jsonify({"status":"full","count":len(favorites)}), 400
    favorites.append(name)
    save_favorites(favorites)
    return jsonify({"status":"added","count":len(favorites)})

@app.route("/api/price/<name>")
def api_price(name):
    korean_name = get_korean_name(name)
    results = crawling(korean_name, limit=5)
    return jsonify({"keyword": korean_name + " 카드", "results": results})

@app.route("/api/crawl-favorites")
def crawl_favorites():
    cache = load_cache()
    pokemon_data = cache["pokemon"]
    favorites = load_favorites()
    result = {}
    for eng in favorites:
        korean = get_korean_name(eng)
        items = crawling(korean, limit=5)
        # 5개 미만이면 "시세 예측 불가"로 채움
        while len(items) < 5:
            items.append({"title": "시세 예측 불가", "price": "시세 예측 불가"})
        result[eng] = {
            "korean_name": korean,
            "items": items[:5],
        }
    return jsonify(result)

@app.route("/api/excel-favorites")
def excel_favorites():
    cache = load_cache()
    favorites = load_favorites()

    if not favorites:
        return jsonify({"error": "즐겨찾기가 없습니다."}), 400

    # 크롤링
    crawl_data = {}
    for eng in favorites:
        korean = get_korean_name(eng)
        items = crawling(korean, limit=5)
        while len(items) < 5:
            items.append({"title": "시세 예측 불가", "price": "시세 예측 불가"})
        crawl_data[eng] = {"korean_name": korean, "items": items[:5]}

    # 엑셀 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "포켓몬 카드 시세"

    # 스타일 정의
    RED   = "E3350D"
    YELLOW= "FFCB05"
    DARK  = "1A1A2E"
    GRAY  = "F5F5F5"
    WHITE = "FFFFFF"
    UNAVAIL = "CCCCCC"

    thin = Side(style='thin', color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_cell(cell, text, bg=RED, fg=WHITE, bold=True, size=11):
        cell.value = text
        cell.font = Font(bold=bold, color=fg, size=size, name="맑은 고딕")
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    def data_cell(cell, text, bg=WHITE, fg="333333", bold=False):
        cell.value = text
        cell.font = Font(bold=bold, color=fg, size=10, name="맑은 고딕")
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border

    # 타이틀 행
    ws.merge_cells("A1:C1")
    title_cell = ws["A1"]
    title_cell.value = f"🎴 포켓몬 카드 중고나라 시세 | {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    title_cell.font = Font(bold=True, color=WHITE, size=14, name="맑은 고딕")
    title_cell.fill = PatternFill("solid", fgColor=DARK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    row = 2
    for eng in favorites:
        if eng not in crawl_data:
            continue
        korean = crawl_data[eng]["korean_name"]
        items  = crawl_data[eng]["items"]

        # 포켓몬 헤더
        ws.merge_cells(f"A{row}:C{row}")
        hdr_cell(ws[f"A{row}"], f"▶  {korean}  ({eng})", bg=RED, fg=YELLOW, size=12)
        ws.row_dimensions[row].height = 28
        row += 1

        # 컬럼 헤더
        hdr_cell(ws[f"A{row}"], "No",    bg="2C2C54", fg=YELLOW)
        hdr_cell(ws[f"B{row}"], "상품명", bg="2C2C54", fg=YELLOW)
        hdr_cell(ws[f"C{row}"], "가격",   bg="2C2C54", fg=YELLOW)
        ws.row_dimensions[row].height = 22
        row += 1

        # 아이템 5개
        for i, item in enumerate(items, 1):
            bg = GRAY if i % 2 == 0 else WHITE
            is_unavail = item["title"] == "시세 예측 불가"
            fg_color = UNAVAIL if is_unavail else "333333"

            data_cell(ws[f"A{row}"], str(i),          bg=bg, fg=fg_color)
            data_cell(ws[f"B{row}"], item["title"],    bg=bg, fg=fg_color)
            data_cell(ws[f"C{row}"], item["price"],    bg=bg, fg=fg_color, bold=(not is_unavail))
            ws.row_dimensions[row].height = 20
            row += 1

        # 여백 행
        ws.merge_cells(f"A{row}:C{row}")
        ws[f"A{row}"].fill = PatternFill("solid", fgColor="F0F0F0")
        ws.row_dimensions[row].height = 10
        row += 1

    # 컬럼 너비
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 18

    # 파일 전송
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    now_str = datetime.now().strftime("%Y%m%d_%H%M")
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"PokePriceReport_{now_str}.xlsx",
    )

app.jinja_env.globals['enumerate'] = enumerate

if __name__ == "__main__":
    app.run(debug=True)